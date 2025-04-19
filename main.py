import logging
import asyncio
from aiogram import Bot, Dispatcher, types, F
from aiogram.client.default import DefaultBotProperties
from aiogram.types import FSInputFile
from aiogram.enums import ParseMode
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.filters import Command
from decouple import config
from openpyxl import Workbook, load_workbook
from datetime import datetime
from io import BytesIO
import os

# Logging
logging.basicConfig(level=logging.DEBUG)  # Изменил уровень логирования на DEBUG для более подробной информации
logger = logging.getLogger(__name__)

# Config
TOKEN = config("TG_TOKEN")
bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())

# Runtime state
user_infos = []  # список словарей: [{username: ..., fio: ...}]
current_question_type = None
current_question = None
current_options = []
user_results = {}
poll_id_to_user = {}
poll_id_to_data = {}

@dp.startup()
async def setup_commands(bot: Bot):
    commands = [
        types.BotCommand(command="start", description="Начать"),
        types.BotCommand(command="poll", description="Следующий вопрос — опрос"),
        types.BotCommand(command="text", description="Следующий вопрос — текстовый"),
        types.BotCommand(command="finish", description="Завершить и получить результаты"),
    ]
    await bot.set_my_commands(commands)

@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    await message.reply("👋 Привет! Пришли Excel-файл (.xlsx) с Telegram ID пользователей ( можно узнать через этого бота: @username_to_id_bot ) в 1-м столбце, а ФИО (необязательно) — во 2-м.")

@dp.message(Command("finish"))
async def finish(message: types.Message):
    if not user_results:
        await message.reply("❌ Ответов нет.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["Username", "FIO", "Question", "Answer", "Timestamp"])

    # Debug: Print user_results to ensure it's populated correctly
    logger.debug(f"user_results before saving: {user_results}")

    for username, answers in user_results.items():
        for answer in answers:
            question, response, timestamp = answer
            fio = next((user["fio"] for user in user_infos if user["username"] == username), "Unknown")
            ws.append([username, fio, question, response, timestamp])

    file_path = "/tmp/results.xlsx"
    wb.save(file_path)

    await bot.send_document(
        chat_id=message.chat.id,
        document=FSInputFile(file_path, filename="results.xlsx"),
        caption="📊 Итоги опроса"
    )
    os.remove(file_path)

@dp.message(F.document)
async def handle_excel(message: types.Message):
    global user_infos
    doc = message.document
    if not doc.file_name.endswith(".xlsx"):
        await message.reply("❌ Пожалуйста, пришли .xlsx файл.")
        return

    file = await bot.download(doc)
    wb = load_workbook(file)
    ws = wb.active

    user_infos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        username = str(row[0]).strip().lstrip("@") if row[0] else None
        fio = str(row[1]).strip() if len(row) > 1 and row[1] else None
        if username:
            user_infos.append({"username": username, "fio": fio})

    await message.reply(f"✅ Загружено {len(user_infos)} пользователей.\n\nТеперь выбери тип вопроса:\n👉 /poll — с вариантами\n👉 /text — письменный ответ")

@dp.message(Command("poll"))
async def set_poll_mode(message: types.Message):
    global current_question_type
    current_question_type = "poll"
    await message.reply("✅ Тип вопроса: <b>опрос</b>. Теперь отправь вопрос в формате:\n\nВопрос: Какой ваш любимый цвет?\nКрасный\nСиний\nЗеленый", parse_mode="HTML")

@dp.message(Command("text"))
async def set_text_mode(message: types.Message):
    global current_question_type
    current_question_type = "text"
    await message.reply("✅ Тип вопроса: <b>текст</b>. Теперь отправь вопрос в формате:\n\n<code>Вопрос: Ваш вопрос</code>", parse_mode="HTML")

@dp.message(F.text)
async def handle_question(message: types.Message):
    global current_question, current_options

    if not user_infos:
        await message.reply("⚠️ Сначала загрузи Excel-файл.")
        return
    if not current_question_type:
        await message.reply("⚠️ Выбери тип вопроса: /poll или /text")
        return

    # Если сообщение начинается с "Вопрос:", это значит, что ты задаешь новый вопрос
    if message.text.startswith("Вопрос:"):
        if current_question_type == "poll":
            lines = message.text.strip().split("\n")
            if len(lines) < 3 or not lines[0].strip().endswith("?"):
                await message.reply("❌ Формат:\nВопрос: Ваш вопрос\nВариант 1\nВариант 2\n...")
                return

            current_question = lines[0].strip()
            current_options = [line.strip() for line in lines[1:] if line.strip()]
            await send_poll_question(message)

        elif current_question_type == "text":
            current_question = "Вопрос: " + message.text.strip()  # Добавляем префикс "вопрос: "
            current_options = []
            await send_text_question(message)
    else:
        # Если вопрос уже был задан, то это ответ на вопрос
        if current_question:
            user = message.from_user.username
            if user not in user_results:
                user_results[user] = []
            user_results[user].append((current_question, message.text.strip(), datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

            # Debug: Log the answer
            logger.debug(f"Saved answer: {current_question} -> {message.text.strip()} for user {user}")
            logger.debug(f"Current user_results: {user_results}")  # Log the user_results state to check for any issues
            await message.reply("✅ Ваш ответ принят!")

async def send_poll_question(message):
    success = 0
    fail = 0
    for info in user_infos:
        username = info["username"]
        try:
            chat = await bot.get_chat(username)
            poll = await bot.send_poll(
                chat_id=chat.id,
                question=current_question,
                options=current_options,
                is_anonymous=False
            )
            poll_id_to_user[poll.poll.id] = username
            poll_id_to_data[poll.poll.id] = (current_question, current_options)
            success += 1
        except Exception as e:
            logger.warning(f"Не отправлен опрос @{username}: {e}")
            fail += 1
    await message.reply(f"✅ Отправлено: {success}\n⚠️ Ошибок: {fail}")

async def send_text_question(message):
    success = 0
    fail = 0
    for info in user_infos:
        username = info["username"]
        fio = info["fio"]
        try:
            chat = await bot.get_chat(username)
            greeting = f"{fio}," if fio else ""
            await bot.send_message(
                chat_id=chat.id,
                text=f"✍️ {greeting} {current_question}",
                parse_mode="HTML"
            )
            if username not in user_results:
                user_results[username] = []
            success += 1
        except Exception as e:
            logger.warning(f"Не отправлен вопрос @{username}: {e}")
            fail += 1
    await message.reply(f"✅ Отправлено: {success}\n⚠️ Ошибок: {fail}")

@dp.poll_answer()
async def handle_poll_answer(poll: types.PollAnswer):
    username = poll_id_to_user.get(poll.poll_id, "Unknown")
    question, options = poll_id_to_data.get(poll.poll_id, ("Unknown", []))
    answer = options[poll.option_ids[0]] if poll.option_ids else "Без ответа"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if username not in user_results:
        user_results[username] = []

    user_results[username].append((question, answer, timestamp))
    logger.info(f"{username} → '{answer}' на '{question}'")

async def main():
    logger.info("Bot is starting...")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
